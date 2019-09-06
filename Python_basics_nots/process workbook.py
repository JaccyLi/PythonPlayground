import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet0']

    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row, 2)
        corrected_price = cell.value * -1.9
        corrected_cell = sheet.cell(row, 3)
        corrected_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=1,
                       max_row=sheet.max_row,
                       min_col=3,
                       max_col=3)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e1')
    wb.save(filename)

process_workbook('transactions.xlsx')